from sqlalchemy import (
    create_engine,
    Integer,
    Text,
    Column,
    Float,
    DateTime,
    String,
    Date,
    text,
    Boolean,
)
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook, Workbook

# working with database with the help of SQLAlchemy
engine = create_engine("sqlite:///test.db")
print(engine)

Session = sessionmaker(bind=engine)
db = Session()
Base = declarative_base()


class WorkPackage(Base):
    __tablename__ = "work_packages"
    id = Column(Integer, primary_key=True)
    type_id = Column(Integer, nullable=True)
    project_id = Column(Integer, nullable=True)
    subject = Column(String, nullable=True)
    description = Column(Text)
    due_date = Column(Date)
    category_id = Column(Integer, nullable=True)
    status_id = Column(Integer, nullable=True)
    assigned_to_id = Column(Integer, nullable=True)
    priority_id = Column(Integer, server_default=text("0"), nullable=True)
    version_id = Column(Integer)
    author_id = Column(Integer, nullable=True)
    lock_version = Column(Integer, nullable=True)
    done_ratio = Column(Integer, nullable=True)
    estimated_hours = Column(Float(53), nullable=True)
    created_at = Column(DateTime, nullable=True)
    updated_at = Column(DateTime, nullable=True)
    start_date = Column(Date)
    responsible_id = Column(Integer, nullable=True)
    budget_id = Column(Integer, nullable=True)
    position = Column(Integer, nullable=True)
    story_points = Column(Integer, nullable=True)
    remaining_hours = Column(Float(53), nullable=True)
    derived_estimated_hours = Column(Float(53), nullable=True)
    schedule_manually = Column(Boolean, nullable=True)


# Create table
# Base.metadata.create_all(engine)

# ws = load_workbook(filename="work_packages_v2.1.xlsx")
# sheet_object = ws.active

# title of the sheet
# print(sheet_object.title)
# # max rows and columns
# print("Max rows are:", sheet_object.max_row, "Max columns are:", sheet_object.max_column)
# # min rows and columns
# print("Min row is:", sheet_object.min_row, "Min column is:", sheet_object.min_column)
#
# # Access all the columns names of the specific cell using row and column value
# print(sheet_object.cell(row=2, column=4).value)
# print("-----------")

# Access all the column name
# maximum_column = sheet_object.max_column
# for col in range(1, maximum_column + 1):
#     print(sheet_object.cell(row=1, column=col).value)
#
# print("-----------")
# # Access all the value of the specific columns
# maximum_rows = sheet_object.max_row
# for row in range(1, maximum_rows + 1):
#     print(sheet_object.cell(row=row, column=4).value)
#
# print("---------------------------")

# Convert each object into python dictionary
# lst = []
# data = {}
# for row in range(1, maximum_rows+1):
#     for col in range(1, maximum_column+1):
#         key = sheet_object.cell(row=1, column=col).value
#         value = sheet_object.cell(row=row, column=col).value
#         data.update({key: value})
#     lst.append(data.copy())
#     data.clear()
#     print()


# for i in lst:
#     print(i)

# print("---------------------------")
# # Append all the value of each cell into the list
# print(sheet_object.dimensions)
# # Produce cells by rows
# x = sheet_object.iter_rows()
# for i in x:
#     l = []
#     for data in i:
#         l.append(data.value)
#     print(l)


# print("---------------------------")
# # Append all the value of each cell into the list
# print(sheet_object.dimensions)
# # Produce cells by cols
# x = sheet_object.iter_cols()
# for i in x:
#     l = []
#     for data in i:
#         l.append(data.value)
#     print(l)

# find the parent of the sheet_object
# print(sheet_object.parent)

# Print the mime_type of the sheet object
# print(sheet_object.mime_type)


# Write excel file using Python openpyxl
wb = Workbook()
ws1 = wb.active
ws1.title = "Students Data"
print(ws1.title)


c1 = ws1.cell(row=1, column=1)
c1.value = "Vishvajit"

c2 = ws1.cell(row=1, column=2)
c2.value = "Rao"

c3 = ws1["A2"]
c3.value = "Alex"

c4 = ws1["B2"]
c4.value = "Doe"


# name of the all sheets in workbook and return list
print(wb.sheetnames)

# access second cell value of first column
print(ws1.cell(row=2, column=1).value)

# access second cell value of second column
print(ws1["B2"].value)

print("------------------------")
# Access all the column's value
for i in ws1.columns:
    for data in i:
        print(data.value)

# wb.save("students.xlsx")